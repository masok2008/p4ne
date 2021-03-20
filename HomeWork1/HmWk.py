# HW1 Py file

import ipaddress
import re
import glob
import openpyxl


def putipplantoxls(fname: str, netlist: list):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title="Address_plan")
    ws['A1'] = "Network"
    ws['B1'] = "Prefix"
    ws['C1'] = "Netmask"
    strnum = 2
    for i in netlist:
        ws['A' + str(strnum)] = str(i.network_address)
        ws['B' + str(strnum)] = "/" + str(i.prefixlen)
        ws['C' + str(strnum)] = str(i.netmask)
        strnum += 1
    wb.save(filename=fname)


def classify(line: str):
    m = re.search(r'^\s*ip address\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)\s*$', line)
    if bool(m):
        ipidr = m.group(1) + "/" + m.group(2)
        #        print(ipidr)
        ip = ipaddress.IPv4Interface(ipidr)
        return {'ip': ip}
    m = re.search(r'^\s*interface\s+(\S+)\s*', line)
    if bool(m):
        return {'int': m.group(1)}
    m = re.search(r'^\s*hostname\s+(\S+)\s*', line)
    if bool(m):
        return {'host': m.group(1)}
    return {}


def findallipint(pathtodir: str):
    netobj = dict()
    netobj['ip'] = []
    netobj['int'] = []
    netobj['host'] = []
    files = glob.glob(pathtodir + "\\*.txt")
    for fn in files:
        with open(fn) as f:
            for st in f:
                d = classify(st)
                for key, value in d.items():
                    netobj[key].append(value)

    netobj['ip'] = list(set(netobj['ip']))
    netobj['int'] = list(set(netobj['int']))
    netobj['host'] = sorted(list(set(netobj['host'])))
    return netobj['ip']


def netlistgen(ipintlist):
    netlst = set()
    for ipidr in ipintlist:
        netlst.add(ipidr.network)
    return list(netlst)


pathtofiles = "C:\\Users\\user\\Seafile\\p4ne_training\\config_files"
fname='TestAddrPlan.xlsx'
ipints = findallipint(pathtofiles)
# for i in sorted(ipints):
#    print(i, end=' ')
# print("")
netlist = netlistgen(ipints)
netlist.sort()
# for i in netlist:
#    print(i, end=' ')
# print("")
for i in netlist:
    print("net: ", i.network_address, "mask:", i.netmask, "(/" + str(i.prefixlen) + ")")
print("")
putipplantoxls(fname, netlist)